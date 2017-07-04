# -*- coding: utf-8 -*-

from django.contrib.auth.models import User
from app import models
from app.models import *
from django.shortcuts import get_object_or_404
from django.conf import settings
import pytz
import os
from openpyxl.reader.excel import load_workbook
from app.utils import *
from datetime import datetime  # 请直接使用datetime()函数


def auth_user(form):  # 瞎写的东西
    if form.is_valid():
        data = form.cleaned_data
        username = data['username']
        password = data['password']
        # check
        user = models.User.objects.filter(username=username, password=password)
        if user:
            return True
        else:
            return False
    else:
        return False


# 提交某个作业（的文件），只能是负责人提交
def submit_homework_file(request):
    workMeta_id = request.POST.get('workMeta_id')
    user = get_object_or_404(User, username=request.user.username)
    enroll = Enroll.objects.get(user=user)
    member = get_object_or_404(Member, user=user)
    print(member.role)
    work = Work(team=member.team, workMeta_id=workMeta_id)
    work.save()
    for file in request.FILES.getlist('files'):
        f = file
        file = File(file=f, course=enroll.course, user_id=user.id)
        file.save()
        Attachment(file=file, work=work, workMeta_id=workMeta_id).save()


# 获得对应团队，对应课程的提交情况，包括已提交和未提交
def get_submittings(team_id, course_id):
    submittings = {'submitted': [], 'unsubmitted': []}
    # 选择最晚提交作业
    last_submit = {}
    for work in Work.objects.filter(team_id=team_id):
        if work.workMeta.submits >= -1:
            if work.workMeta_id not in last_submit:
                last_submit[work.workMeta_id] = {'work': work,
                                                 'time': datetime(2017, 1, 1, tzinfo=pytz.utc)}
            if work.time > last_submit[work.workMeta_id]['time']:
                last_submit[work.workMeta_id]['time'] = work.time
                last_submit[work.workMeta_id]['work'] = work

    last_submit = sorted(last_submit.items(), key=lambda d: d[0], reverse=False)
    submittings['submitted'] = [item[1]['work'] for item in last_submit]
    submitted_id = [item.workMeta_id for item in submittings['submitted']]

    # 没有提交的作业
    for workMeta in WorkMeta.objects.filter(course_id=course_id):
        if workMeta.id not in submitted_id:
            if check_submit_time(workMeta):
                submittings['unsubmitted'].append(workMeta)
    return submittings


def handle_uploaded_contribution(request, f=None):
    # datenow = datetime.datetime.now()
    # filedate = datenow.strftime('%Y%m%d-%H%M%S')
    # path = IMPORT_ROOT
    # filepath = path + '/' + filedate + '_' + f.name
    # with open(filepath, 'ab') as de:
    #     for chunk in f.chunks():
    #         de.write(chunk)
    # wb = load_workbook(filepath)
    # log(filepath)
    # table = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    datenow = datetime.now()
    filedate = datenow.strftime('%Y%m%d-%H%M%S')
    path = UPLOAD_ROOT
    log(path, 'handle_uploaded_user')
    filepath = os.path.join(path, filedate + '_' + f.name)
    log(filepath, 'handle_uploaded_user')
    with open(filepath, 'wb+') as de:
        for chunk in f.chunks():
            de.write(chunk)
    # 导入exel
    wb = load_workbook(filepath)
    log('导入exel', 'handle_uploaded_user')
    table = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    for i in range(2, table.max_row + 1):
        if table.cell(row=i, column=1).value is None:
            # '为空，应跳过'
            continue
        print('正在导入第' + str(i - 1) + '行...')

        student_id = table.cell(row=i, column=1).value
        student_contribution = table.cell(row=i, column=3).value
        print(student_id, student_contribution)
        student = User.objects.filter(username=student_id).first()
        # if student is not None:
        #     Member.objects.filter(user=student).update(contribution=student_contribution)
        # else:
        #     return None
        setContribution(student, student_contribution)


def check_submit_time(workMeta):
    return workMeta.startTime < datetime.now(tz=pytz.utc) < workMeta.endTime
