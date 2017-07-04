from app.utils.logUtils import log, LOG_LEVEL
from app.utils.authUtils import isStudent
from datetime import datetime
from app.models import *
from openpyxl import *

# 关于签到的工具集
# by kahsolt


def showToday():
    startTime = datetime(datetime.now().year, datetime.now().month, datetime.now().day, 9, 0, 0, 0)
    endTime = datetime.now()
    attendances = Attendance.objects.filter(time__range=(startTime, endTime)).order_by('time')
    return _distinct(attendances)


def showTimeBetween(startTime, endTime):
    if not isinstance(startTime, datetime) or not isinstance(endTime, datetime):
        return None

    attendances = Attendance.objects.filter(time__range=(startTime, endTime)).order_by('time')
    return _distinct(attendances)


def _distinct(atts):
    ret = []
    u = []
    for att in atts:
        if not att.user in u:
            u.append(att.user)
            ret.append(att)
    return ret


def addAttendance(user):
    if not isinstance(user, User):
        return None
    if user.role != 'student':
        log('不是学生', 'attendanceUtils', LOG_LEVEL.ERROR)
        return False

    attendence = Attendance()
    attendence.user = user
    attendence.save()
    return attendence


def writeAttendanceReport(file, course_id, attendance=None):
    if attendance is None:
        attendance = showToday()

    attendance_stu = [att.user for att in attendance]
    work_book = Workbook()
    ws = work_book.get_active_sheet()
    ws.cell(row=1, column=1).value = '学号'
    ws.cell(row=1, column=2).value = '姓名'
    ws.cell(row=1, column=3).value = '签到'

    num = 2
    users = Enroll.objects.filter(course_id=course_id)
    for user in users:
        if isStudent(user.user):
            ws.cell(row=num, column=1).value = str(user.user.username)
            ws.cell(row=num, column=2).value = str(user.user.name)
            if user in attendance_stu:
                ws.cell(row=num, column=3).value = '1'
            else:
                ws.cell(row=num, column=3).value = '0'
            num += 1
    work_book.save(filename=file)
