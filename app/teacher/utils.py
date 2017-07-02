# -*- coding: utf-8 -*-
# define your utility function here
import os
from datetime import datetime
from openpyxl import *
from app.models import *
from django.shortcuts import get_object_or_404
from app.utils import *
from app.utils.logUtils import *
from app.teacher.entities import *
from django.core.exceptions import *


def handle_uploaded_file(request,course_id,f):
    # f = request.FILES['file']
    # name = f.name
    # new_name = name.split('.', 1)[0] + '-' + datetime.now().strftime("%Y-%m-%d-%H-%M-%S").__str__() + '.' + \
    #            name.split('.', 1)[1]
    # with open(os.path.join('uploads/file/teacher/', new_name), 'wb+') as destination:
    #     for chunk in f.chunks():
    #         destination.write(chunk)
    #f=form.cleaned_data['file']
    teacher=get_object_or_404(User,username=request.user.username)
    file=File(course_id=course_id,time=datetime.now(),file=f,user=teacher)
    file.save()


def import_student_for_course(request):
    course_id = request.POST.get('course_id', None)
    f = request.FILES['file']
    name = f.name
    if isXls(name):
        path = os.path.join(IMPORT_ROOT, name)
        with open(path, 'wb+') as destination:
            for chunk in f.chunks():
                destination.write(chunk)
        wb = load_workbook(filename=path)
        sheet = wb.active
        ids = []
        for i in range(2, 10000):
            if sheet['A' + str(i)] is None:
                break
            else:
                ids.append(sheet['A' + str(i)])
        students_to_course(ids, course_id)
        os.remove(path)
        return 'import students for course success'
    else:
        return 'upload file must be a xls or xlsx excel'


def students_to_course(students_id, course_id):
    User.objects.filter(role='student', course_id=course_id).delete()
    course = Course.objects.get(id=course_id)
    for id in students_id:
        student = User.objects.get(username=id)
        enroll = Enroll(course=course, user=student)
        enroll.save()


def add_homework(homework_form, course_id, username,file):
    content = homework_form.cleaned_data['content']
    proportion = homework_form.cleaned_data['proportion']
    submits = homework_form.cleaned_data['submits']
    endTime = homework_form.cleaned_data['endTime']
    title=homework_form.cleaned_data['title']
    teacher = get_object_or_404(User, username=username)
    workmeta = WorkMeta(course_id=course_id, user=teacher, content=content,title=title,
                        proportion=proportion, submits=submits, startTime=datetime.now(), endTime=endTime)
    workmeta.save()
    if file is not None:
        f=File(course_id=course_id,user=teacher,file=file,type='text',time=datetime.now())
        f.save()
        attachment = Attachment(file=f, workMeta=workmeta, type='workmeta')
        attachment.save()


# 获取一个老师往期课程的所有作业
def get_past_homeworks(username):
    teacher = get_object_or_404(User, username=username)
    enrolls = Enroll.objects.filter(user__username=teacher.username)
    homeworks = []
    present = datetime.now()
    for enroll in enrolls:
        course = enroll.course
        if course.endTime.replace(tzinfo=None) <= present:
            workmetas = WorkMeta.objects.filter(course_id=course.id)
            homeworks.extend(workmetas)
    return homeworks

#获取当前学期
def get_now_term():
    now_term = Term.objects.all().order_by('-id')[0]
    return now_term


# 获取团队成绩表的完整路径名
def get_team_score_excel_file_abspath():
    now_term = get_now_term()

    # 第一次计算出各团队得分之后保存到excel表，以便老师下载
    # 各个团队得分的excel命名规范：termYear_termsemester_team_score_list.xlsx
    file_path = os.path.join(REPORT_ROOT, 'teamScores')
    if not os.path.exists(file_path):
        os.mkdir(file_path,)
    file_name = '' + str(now_term.year) + str(now_term.semester) + '_team_score_list.xlsx'
    file = os.path.join(file_path, file_name)
    return file


# 获取学生成绩表的excel的完整路径名
def get_stu_score_excel_file_abspath():
    now_term = get_now_term()

    # 各人得分的excel命名规范：termYear_termsemester_stu_score_list.xlsx
    file_path = os.path.join(REPORT_ROOT, 'stuScores')
    if not os.path.exists(file_path):
        os.mkdir(file_path,)
    file_name = '' + str(now_term.year) + str(now_term.semester) + '_stu_score_list.xlsx'
    file = os.path.join(file_path, file_name)
    return file

# 获取所有的团队学生表的excel的完整路径名
def get_team_members_all_excel_file_abspath():
    now_term = get_now_term()

    # 团队成员表的excel命名规范：termYear_termsemester_stu_teams.xlsx
    file_path = os.path.join(REPORT_ROOT, 'stuTeams')
    file_name = '' + str(now_term.year) + str(now_term.semester) + '_stu_teams.xlsx'
    file = os.path.join(file_path, file_name)
    return file

# 保存所有的学生和团队信息到excel
def create_stu_teams_excel(file,course):
    work_book = Workbook()
    team_stu_list = reportTeams(course)
    ws = work_book.get_active_sheet()
    print(team_stu_list)

    ws.cell(row=1, column=1).value = '团队id'
    ws.cell(row=1, column=2).value = '团队名称'
    ws.cell(row=1, column=3).value = '队长'
    ws.cell(row=1,column=4).value = '队员'

    for i in range(0,len(team_stu_list)):
        num = i + 2
        ws.cell(row=num, column=1).value = team_stu_list[i]['id']
        ws.cell(row=num, column=2).value = team_stu_list[i]['name']
        ws.cell(row=num, column=3).value = team_stu_list[i]['leader'].user.name
        column_index = 4
        for member in team_stu_list[i]['member']:
            ws.cell(row=num, column=column_index).value = member.user.name
            column_index += 1
        work_book.save(filename=file)


# 保存团队得分表到excel
def create_team_score_excel(file, team_list, score_list):
    work_book = Workbook()
    ws = work_book.get_active_sheet()
    ws.cell(row=1, column=1).value = '团队id'
    ws.cell(row=1, column=2).value = '团队名称'
    ws.cell(row=1, column=3).value = '分数'

    for i in range(0,len(team_list)):
        num = i+2;
        ws.cell(row=num, column=1).value = team_list[i].serialNum
        ws.cell(row=num, column=2).value = team_list[i].name
        ws.cell(row=num, column=3).value = score_list[i]
    work_book.save(filename=file)


def create_stu_score_excel(file, stu_list, stu_score_dict):
    work_book = Workbook()
    ws = work_book.get_active_sheet()
    ws.cell(row=1, column=1).value = '学号'
    ws.cell(row=1, column=2).value = '姓名'
    ws.cell(row=1, column=3).value = '分数'

    print(stu_list)
    print(stu_score_dict)
    num = 2
    for stu in stu_list:
        ws.cell(row=num, column=1).value = stu.username
        ws.cell(row=num, column=2).value = stu.name
        ws.cell(row=num, column=3).value = stu_score_dict[stu]
        num += 1
    work_book.save(filename=file)



#计算团队得分
def compute_team_score():
    # now_term = get_now_term()
    team_list = get_team_list_in_now_course()
    score_list = []
    team_score = {}
    for team in team_list:
        work_list = Work.objects.filter(team=team)
        score = 0
        for work in work_list:
            score += (work.score or 0.0) * (work.workMeta.proportion or 0.0)
        team_score[team] = score
        score_list.append(score)
    return team_list, score_list, team_score


# 计算个人总得分 dict: key=team_member, value=score
def compute_stu_score():
    x, y, team_score = compute_team_score()
    stu_score = {}
    for team in team_score:
        team_member = Member.objects.filter(team=team)
        for member in team_member:
            score = (team_score[team] or 0.0) * (member.contribution or 0.0)
            stu_score[member.user] = score
    return stu_score



# 获取当前学期的所有team,并排序
def get_team_list_in_now_course():
    now_time = datetime.now()
    now_course = Course.objects.get(startTime__lt=now_time, endTime__gt=now_time)
    if now_course:
        team_list = Team.objects.filter(course=now_course).order_by('serialNum')
    return team_list


# 获取当前学期所有参加的学生，并按学号排序
def get_stu_list_in_now_course():
    team_list = get_team_list_in_now_course()
    members_list = []
    for team in team_list:
        team_member = Member.objects.filter(team=team)
        members_list.extend(team_member)
    stu_list = []
    for member in members_list:
        stu_list.append(member.user)
    stu_list = sorted(stu_list, key=lambda x : x.username)
    return stu_list


# 读取file文件
def file_iterator(file_name, chunk_size=512):
    with open(file_name, 'rb') as f:
        while True:
            c = f.read(chunk_size)
            if c:
                yield c
            else:
                break


# 获取某一门课程中没有团队的学生
def query_unteamed_students(course_id):
    course=get_object_or_404(Course, id=course_id)
    enrolls=Enroll.objects.filter(course=course,user__role='student')
    unteamed_students=[]
    for enroll in enrolls:
        try:
            member=Member.objects.get(user=enroll.user)
        except:
            unteamed_students.append(enroll.user)
    return unteamed_students


# 生成所有团队的所有成绩报表
def generate_team_scores(course_id):
    course=get_object_or_404(Course,id=course_id)
    workmetas=WorkMeta.objects.filter(course=course)
    teams=Team.objects.filter(course=course)
    datas=[]
    for workmeta in workmetas:
        works=[]
        for team in teams:
            work=Work.objects.filter(workMeta=workmeta,team=team).order_by('-time').first()
            if work is None:
                work=Work(score=0,team=team)
            works.append(work)
        row_data=ScoreWrapper(workmeta=workmeta,works=works)
        datas.append(row_data)
    return datas


def generate_scores_excel(course_id):
    data=generate_team_scores(course_id)
    wb=Workbook()
    dest=os.path.join(REPORT_ROOT,'teams_scores_'+str(course_id)+'.xlsx')
    ws1=wb.active
    ws1.title='团队成绩报表'
    ws1['A1']='作业标题\团队'
    ws1['A'+str(len(data)+3)]='加权总成绩'
    teams=Team.objects.filter(course_id=course_id)

    for i in range(len(teams)):
        ws1.cell(row=1,column=i+2,value=teams[i].name)

    for i in range(len(data)):
        ws1.cell(row=i+2,column=1,value=data[i].workmeta.title)
        works=data[i].works
        for j in range(len(works)):
            ws1.cell(row=i+2,column=j+2,value=works[j].score)

    # 根据每次作业的权重，计算总成绩
    for i in range(len(teams)):
        sum=0
        for j in range(len(data)):
            weight=data[j].workmeta.proportion
            sum+=weight*data[j].works[i].score
        ws1.cell(row=len(data)+3,column=i+2,value=sum)
    wb.save(filename=dest)
    return dest
