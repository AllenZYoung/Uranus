# -*- coding: utf-8 -*-
#define your utility function here

from datetime import datetime
import os
from app import models
from django.contrib.auth.models import User
from openpyxl.reader.excel import load_workbook
from Uranus.settings import  BASE_DIR
from app.utils import log
#from app.utils import *

def getItem(user_id):
    try :
        user = models.User.objects.get(username=user_id)
        return user
    except Exception as e:
        return None


def handle_uploaded_user(request,course_id, f=None, user_role='student'):
    # 向resource中传入上传的exel
    datenow = datetime.now()
    filedate = datenow.strftime('%Y%m%d-%H%M%S')
    path = os.path.join(BASE_DIR,'resource','uploads','import')
    log(path, 'handle_uploaded_user')
    filepath = os.path.join(path,filedate+'_'+f.name)
    log(filepath, 'handle_uploaded_user')
    with open(filepath, 'wb+') as de:
        for chunk in f.chunks():
            de.write(chunk)
    #导入exel
    wb = load_workbook(filepath)
    log('导入exel', 'handle_uploaded_user')
    table = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    #如果是重新导入删除原有enroll
    enrolls = models.Enroll.objects.filter(course=course_id)
    if enrolls:
        for enroll in enrolls:
            user  = enroll.user
            if user.role == user_role:
                enroll.delete()
    #创建enroll
    exception_list = []
    for i in range(2, table.max_row + 1):
        if table.cell(row=i, column=1).value is None:
            # 为空，应跳过'
            continue
        print('正在导入第' + str(i - 1) + '行...')
        user_id = table.cell(row=i,column=1).value
        user_sex = table.cell(row=i, column=3).value
        sex = 'male'
        if (user_sex.find('女') != -1):
            sex = 'female'
        #数据库已存在相同的用户信息，则不创建用户只创建enroll
        #如果不存在用户，则创建用户
        get_user = getItem(user_id)
        if get_user == None:
            model_user = models.User(
                username=user_id,
                password='User123',
                name=table.cell(row=i, column=2).value,
                role=user_role,
                gender=sex,
                tel=table.cell(row=i, column=4).value
            )
            if (user_role == 'student'):
                model_user.classID = classID = table.cell(row=i, column=5).value
            model_user.save()
            model_enroll = models.Enroll(
                course=models.Course.objects.get(id=course_id),
                user=model_user
            )
            model_enroll.save()
            User.objects.create_user(username=user_id, password='User123')
        else:
            if get_user.role == user_role :
                model_enroll = models.Enroll(
                    course=models.Course.objects.get(id=course_id),
                    user=get_user
                )
                model_enroll.save()
            else:
                if user_role == 'student':
                    exception_list.append('第'+str(i)+'行的学生学号与已存在的老师相同')
                else:
                    exception_list.append('第' + str(i) + '行的老师工号与已存在的学生相同')
    if exception_list :
        exception_message = ''.join(exception_list)
        raise Exception(exception_message)